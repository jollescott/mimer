import os
import json
import random
from openpyxl import load_workbook


def scrape():
    wb = load_workbook('./greenlandic-swedish.xlsx', read_only=True)
    ws = wb.active

    islandic = []
    english = []
    tags = []

    for row in ws.iter_rows(min_row=2):
        original = row[0]
        translated = row[1]
        tag = row[2]

        if original.value is not None and translated.value is not None:

            if len(translated.value.split()) == 1:
                islandic.append(original.value)
                english.append(translated.value)
                tags.append(tag.value)

    return (islandic, english, tags)


def format_questions(words):
    questions = []
    index = 0

    for question in words[0]:
        translation = words[1][index]
        tags = words[2][index]

        question = {
            'greenlandic': question,
            'swedish': translation,
            'tags': tags
        }

        questions.append(question)

        index += 1

    return questions


if __name__ == "__main__":
    if os.path.isfile('./greenlandic-swedish.xlsx'):
        questions = scrape()
        questions = format_questions(questions)

        file = open('questions.json', 'w')
        file.write(json.dumps(questions))
        file.close()
    else:
        print('Dictionary not found.')
